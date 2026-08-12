#!/usr/bin/env python3
"""Regenerate the app's scored database (data/curated_fungi_both.csv) from the matrix.

Reads the per-protein identity matrix and, for each of the six functional categories,
scores each organism from its best hit using the **>= rule**:
    I >= 75  -> 2      I >= 35  -> 1      else -> 0
Names are cleaned (underscores -> spaces, whitespace collapsed). Output columns match
the app's DB: Phyla, Species, then the six 0/1/2 category scores.

After running this, rebuild the embedded data with:  python3 build_data.py

Usage: python3 build_curated.py [MATRIX.xlsx] [-o data/curated_fungi_both.csv]
"""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import openpyxl

DEFAULT_MATRIX = ("/Users/aam/GitHub/Predicting-Fungal-Contaminants-for-Space-Missions/"
                  "data/Updated_fungi_proteins_fixed.xlsx")
HERE = Path(__file__).parent

# category -> (output column name, query proteins)
CATEGORIES = [
    ("antimicrobial-resistance", ["ERG11", "CDR1", "MDR1", "UPC2", "TAC1", "MRR1"]),
    ("Biofilm-formation",        ["BCR1", "EFG1", "TEC1", "HWP1", "ALS3", "NDT80", "ERG251", "CZF1", "FLO8"]),
    ("Human-pathogenicity",      ["SAP5", "PLB1", "LAC1", "RIM101"]),
    ("Thermophile",              ["HSP90"]),
    ("Radiation-resistance",     ["RAD51"]),
    ("Spore-formation",          ["brlA", "abaA", "wetA", "srr1"]),
]
COLS = [c for c, _ in CATEGORIES]


def clean(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).replace("_", " ")).strip()


def num(x):
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    if "," in s:
        try:
            return max(float(p) for p in s.split(",") if p.strip())
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def score(vals):
    m = max([v for v in vals if v is not None], default=None)
    return 2 if (m is not None and m >= 75) else 1 if (m is not None and m >= 35) else 0


def build(matrix, out):
    wb = openpyxl.load_workbook(matrix, data_only=True, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    hdr = [str(h) if h is not None else "" for h in rows[1]]
    col = {p: hdr.index(p) for _, gs in CATEGORIES for p in gs if p in hdr}
    pidx = next((i for i, h in enumerate(hdr) if h.strip().lower() in ("phyla", "phylum")),
                len(hdr) - 1)
    out_rows = []
    for r in rows[2:]:
        if not r[0] or not str(r[0]).strip() or not r[pidx] or not str(r[pidx]).strip():
            continue
        scores = [score([num(r[col[g]]) for g in gs]) for _, gs in CATEGORIES]
        out_rows.append([str(r[pidx]).strip(), clean(r[0])] + scores)

    with open(out, "w", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["Phyla", "Species"] + COLS)
        w.writerows(out_rows)
    print(f"wrote {out}: {len(out_rows)} organisms (>= rule, cleaned names)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("matrix", nargs="?", default=DEFAULT_MATRIX)
    ap.add_argument("-o", "--out", default=str(HERE / "data" / "curated_fungi_both.csv"))
    args = ap.parse_args()
    build(args.matrix, args.out)
